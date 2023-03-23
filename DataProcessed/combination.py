import fangzi_strcut
import openpyxl

"""
    将path2的方子更新到path下的表格中
"""

path = r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\别名替换后的方子1-4.xlsx'
sheet_name = '方子'
excel = openpyxl.load_workbook(path)
sheet = excel[sheet_name]

path2 = r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\1_含有无剂量单位药材的方子-zb(1).xlsx'
sheet_name2 = '方子'
sheet2, fangzi_object2 = fangzi_strcut.open_excel(path2, sheet_name2)

# 保存path2下的所有方子id  key: fangzi.id  value： fangzi.herb_total
fangzi_id = {}
k = 0
for fangzi in fangzi_object2:
    if k == 0:
        k = 1
        continue
    fangzi_id[fangzi.id] = fangzi.herb_total

for i in range(sheet.max_row):
    fangzi = fangzi_strcut.Struct_of_fangzi(sheet, i)
    if fangzi.id in fangzi_id.keys():
        sheet.cell(i + 1, 6, fangzi_id[fangzi.id])


excel.save(r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\别名替换后的方子1-5.xlsx')