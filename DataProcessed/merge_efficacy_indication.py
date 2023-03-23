import fangzi_strcut
import openpyxl

"""
    程序描述：
	将功效一行中的数据直接加到同一行的主治中，若两者都为”-“则将此方子删除
	结果保存在”65776方子.txt“中.此时方子总数为65776.
"""

path = r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\65788方子.xlsx'
sheet_name = '方子'
sheet, fangzi_object = fangzi_strcut.open_excel(path, sheet_name)

excel1 = openpyxl.Workbook()
sheet1 = excel1.create_sheet('方子')


def copy_sheet(sheet, sheet2, i, k, column_11, column_12):
    sheet2.cell(k, 1, sheet.cell(i + 1, 1).value)
    sheet2.cell(k, 2, sheet.cell(i + 1, 2).value)
    sheet2.cell(k, 3, sheet.cell(i + 1, 3).value)
    sheet2.cell(k, 4, sheet.cell(i + 1, 4).value)
    sheet2.cell(k, 5, sheet.cell(i + 1, 5).value)
    sheet2.cell(k, 6, sheet.cell(i + 1, 6).value)
    sheet2.cell(k, 7, sheet.cell(i + 1, 7).value)
    sheet2.cell(k, 8, sheet.cell(i + 1, 8).value)
    sheet2.cell(k, 9, sheet.cell(i + 1, 9).value)
    sheet2.cell(k, 10, sheet.cell(i + 1, 10).value)

    sheet2.cell(k, 11, column_11)
    sheet2.cell(k, 12, column_12)

    sheet2.cell(k, 13, sheet.cell(i + 1, 13).value)
    sheet2.cell(k, 14, sheet.cell(i + 1, 14).value)
    sheet2.cell(k, 15, sheet.cell(i + 1, 15).value)
    sheet2.cell(k, 16, sheet.cell(i + 1, 16).value)
    sheet2.cell(k, 17, sheet.cell(i + 1, 17).value)
    sheet2.cell(k, 18, sheet.cell(i + 1, 18).value)
    sheet2.cell(k, 19, sheet.cell(i + 1, 19).value)
    sheet2.cell(k, 20, sheet.cell(i + 1, 20).value)

k = 1

for i in range(len(fangzi_object)):
    # 第一行抬头都要复制过去
    if i == 0:
        copy_sheet(sheet, sheet1, i, k, sheet.cell(i + 1, 11).value, sheet.cell(i + 1, 12).value)
        k = k + 1
        continue

    efficacy = fangzi_object[i].efficacy
    indication = fangzi_object[i].indication

    if indication != "-":
        copy_sheet(sheet, sheet1, i, k, "-", indication)
        k += 1


excel1.save(r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\63702方子.xlsx')