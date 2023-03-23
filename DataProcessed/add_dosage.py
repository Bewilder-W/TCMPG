from fangzi_strcut import Struct_of_fangzi
import openpyxl
import re

"""
    将方子中含有各等分的方子进行替换
"""
path = r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\别名替换后的方子1-3.xlsx'
sheet_name = '方子'
excel = openpyxl.load_workbook(path)
sheet = excel[sheet_name]

# # 保存所有方子的对象
# fangzi_object = []

def split_column_FGH(string_of_cellvalue):
    result = string_of_cellvalue.split(' ')
    result = [i for i in result if i != '']
    return result


for i in range(sheet.max_row):
    fangzi = Struct_of_fangzi(sheet, i)

    herb_total = split_column_FGH(fangzi.herb_total)
    herb_non_unit = split_column_FGH(fangzi.herb_non_unit)

    flag = 0
    for m, n in zip(herb_total, herb_non_unit):
        k = str(m.replace(str(n), ''))
        if k != '':
            flag = 1
    if flag == 0:
        new_herb_total = "" + herb_non_unit[0] + "3两"
        for k in range(len(herb_non_unit)):
            if k == 0:
                continue
            new_herb_total += " " + herb_non_unit[k] + "3两"
        # print(fangzi.id + new_herb_total)
        sheet.cell(i+1, 6, new_herb_total)

    # if not bool(re.search(r'\d', herb_total)):
    #     if "各等分" in herb_total or "等分" in herb_total:
    #         herb_non_unit = split_column_FGH(fangzi.herb_non_unit)
    #         new_herb_total = "" + herb_non_unit[0] + "3两"
    #         for j in range(len(herb_non_unit)):
    #             if j == 0:
    #                 continue
    #             new_herb_total += " " + herb_non_unit[j] + "3两"
    #         sheet.cell(i+1, 6, new_herb_total)
excel.save(r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\别名替换后的方子1-4.xlsx')