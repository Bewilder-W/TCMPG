import openpyxl
import fangzi_strcut
"""
    程序描述：
        以"84463方子-10000.xlsx"表中的data子表为操作对象，新建两个excel表格，分别
        为"列FGH药材数量不等的方子.xlsx"和"列FGH药材数量相等的方子.xlsx"。遍历每一
        个方子的时候，判断列F在standard_herbtotal（）处理后的列表长度和列G在
        split_column_GH（）处理后的列表长度是否一样。若数量相等，则保存于
        "列FGH药材数量相等的方子.xlsx"，否则保存于"列FGH药材数量不等的方子.xlsx"
    
    程序输入：84463方子-10000.xlsx
    程序输出：
        C:\\Users\吴杨\Desktop\中药复方-吴杨\TCM\列FGH药材数量不等的方子.xlsx
        C:\\Users\吴杨\Desktop\中药复方-吴杨\TCM\列FGH药材数量相等的方子.xlsx
"""

path = r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\84463方子-10000.xlsx'
sheet_name = 'data'
sheet, fangzi_object = fangzi_strcut.open_excel(path, sheet_name)

# 将sheet的第i+1行复制给sheet2的第K行
def copy_sheet(sheet, sheet2, i, k):
    sheet2.cell(k, 1, sheet.cell(i + 1, 1).value)
    sheet2.cell(k, 2, sheet.cell(i + 1, 2).value)
    sheet2.cell(k, 3, sheet.cell(i + 1, 3).value)
    sheet2.cell(k, 4, sheet.cell(i + 1, 4).value)
    sheet2.cell(k, 5, sheet.cell(i + 1, 5).value)
    sheet2.cell(k, 6, sheet.cell(i + 1, 6).value)
    sheet2.cell(k, 7, sheet.cell(i + 1, 7).value)
    sheet2.cell(k, 8, sheet.cell(i + 1, 8).value)
    sheet2.cell(k, 9, sheet.cell(i + 1, 9).value)
    sheet2.cell(k, 10, sheet.cell(i + 1, 10).value)
    sheet2.cell(k, 11, sheet.cell(i + 1, 11).value)
    sheet2.cell(k, 12, sheet.cell(i + 1, 12).value)
    sheet2.cell(k, 13, sheet.cell(i + 1, 13).value)
    sheet2.cell(k, 14, sheet.cell(i + 1, 14).value)
    sheet2.cell(k, 15, sheet.cell(i + 1, 15).value)
    sheet2.cell(k, 16, sheet.cell(i + 1, 16).value)
    sheet2.cell(k, 17, sheet.cell(i + 1, 17).value)
    sheet2.cell(k, 18, sheet.cell(i + 1, 18).value)
    sheet2.cell(k, 19, sheet.cell(i + 1, 19).value)
    sheet2.cell(k, 20, sheet.cell(i + 1, 20).value)

# 新建一个excel表格保存列FGH药材数量不等的方子
excel1 = openpyxl.Workbook()
sheet1 = excel1.create_sheet('方子')

# 新建一个excel表格保存列FGH药材数量相等的方子
excel2 = openpyxl.Workbook()
sheet2 = excel2.create_sheet('方子')

k1 = 1  # 记录sheet1的行号
k2 = 1  # 记录sheet2的行号
for i in range(len(fangzi_object)):
    # 第一行抬头都要复制过去
    if i == 0:
        copy_sheet(sheet, sheet1, i, k1)
        k1 = k1 + 1
        copy_sheet(sheet, sheet2, i, k2)
        k2 = k2 + 1
        continue
    herb_total = fangzi_strcut.standard_herbtotal(fangzi_object[i].herb_total)
    herb_non_unit = fangzi_strcut.split_column_FGH(fangzi_object[i].herb_non_unit)

    if len(herb_total) != len(herb_non_unit):
        copy_sheet(sheet, sheet1, i, k1)
        k1 = k1 + 1
    else:
        copy_sheet(sheet, sheet2, i, k2)
        k2 = k2 + 1


excel1.save(r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\列FGH药材数量不等的方子.xlsx')
excel2.save(r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\列FGH药材数量相等的方子.xlsx')