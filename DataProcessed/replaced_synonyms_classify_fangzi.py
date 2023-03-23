import openpyxl
import fangzi_strcut
import herb_synonyms_map

"""
    程序描述：
        以"列FGH药材数量相等的方子.xlsx"为操作对象，新建两个excel表格，分别
        为"别名替换后的方子.xlsx"和"别名无法替换的方子.xlsx"。遍历每一个方子，
        用count记录方子的列H中带【】药材的数量，将列H中带【】的药材找出来，
        找到一个【】，count+1.然后截取出去【】的子串，判断子串是否存在于
        herb_synonyms{}中药别名表里。若存在，则直接替换，count-1，否则留下
        对整个列H替换完，如果count不为0，则表示别名替换操作后，列H还有没被
        替换的药材，需要保存到"别名无法替换的方子.xlsx",否则保存到"别名替换后
        的方子.xlsx"。保存的时候需要注意，将列F和列H的格式一并规格化。
    
    程序输入：列FGH药材数量相等的方子.xlsx
    程序输出：
        C:\\Users\吴杨\Desktop\中药复方-吴杨\TCM\别名替换后的方子.xlsx
        C:\\Users\吴杨\Desktop\中药复方-吴杨\TCM\别名无法替换的方子.xlsx
"""

path = r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\列FGH药材数量相等的方子.xlsx'
sheet_name = '方子'
sheet, fangzi_object = fangzi_strcut.open_excel(path, sheet_name)


# 将sheet的第i+1行复制给sheet2的第K行
def copy_sheet(sheet, sheet2, i, k, column_F , column_H):
    sheet2.cell(k, 1, sheet.cell(i + 1, 1).value)
    sheet2.cell(k, 2, sheet.cell(i + 1, 2).value)
    sheet2.cell(k, 3, sheet.cell(i + 1, 3).value)
    sheet2.cell(k, 4, sheet.cell(i + 1, 4).value)
    sheet2.cell(k, 5, sheet.cell(i + 1, 5).value)

    sheet2.cell(k, 6, column_F)

    sheet2.cell(k, 7, sheet.cell(i + 1, 7).value)

    sheet2.cell(k, 8, column_H)

    sheet2.cell(k, 9, sheet.cell(i + 1, 9).value)
    sheet2.cell(k, 10, sheet.cell(i + 1, 10).value)
    sheet2.cell(k, 11, sheet.cell(i + 1, 11).value)
    sheet2.cell(k, 12, sheet.cell(i + 1, 12).value)
    sheet2.cell(k, 13, sheet.cell(i + 1, 13).value)
    sheet2.cell(k, 14, sheet.cell(i + 1, 14).value)
    sheet2.cell(k, 15, sheet.cell(i + 1, 15).value)
    sheet2.cell(k, 16, sheet.cell(i + 1, 16).value)
    sheet2.cell(k, 17, sheet.cell(i + 1, 17).value)
    sheet2.cell(k, 18, sheet.cell(i + 1, 18).value)
    sheet2.cell(k, 19, sheet.cell(i + 1, 19).value)
    sheet2.cell(k, 20, sheet.cell(i + 1, 20).value)


def list_to_str(herb_list):
    result = ""
    for i in range(len(herb_list)):
        result = result + herb_list[i]
        if i != len(herb_list) - 1:
            result = result + " "

    return result

# 新建一个excel表格保存别名替换后的方子
excel1 = openpyxl.Workbook()
sheet1 = excel1.create_sheet('方子')

# 新建一个excel表格保存别名无法替换的方子
excel2 = openpyxl.Workbook()
sheet2 = excel2.create_sheet('方子')

# 取出herb_synonyms_map.py中的一一映射表
herb_synonyms = herb_synonyms_map.herb_synonyms

k1 = 1  # 记录sheet1的行号
k2 = 1  # 记录sheet2的行号

# 保存中文的左右括号
left_bracket = '【'
right_bracket = '】'

for i in range(len(fangzi_object)):
    # 第一行抬头都要复制过去
    if i == 0:
        copy_sheet(sheet, sheet1, i, k1, sheet.cell(i + 1, 6).value, sheet.cell(i + 1, 8).value)
        k1 = k1 + 1
        copy_sheet(sheet, sheet2, i, k2, sheet.cell(i + 1, 6).value, sheet.cell(i + 1, 8).value)
        k2 = k2 + 1
        continue

    count = 0  # 记录每一个方子的列H中剩余带【】药材的数量
    herb_total = fangzi_strcut.standard_herbtotal(fangzi_object[i].herb_total)  # list数据类型
    herb_total = list_to_str(herb_total)  # str数据类型

    herb_replaced_synonyms = fangzi_object[i].herb_replaced_synonyms
    herb_replaced_synonyms = fangzi_strcut.split_column_FGH(herb_replaced_synonyms)  # list数据类型

    # 将列H中的带【】的药材找出来
    for j in range(len(herb_replaced_synonyms)):
        # 如果这个方子中存在【】的药材
        if herb_replaced_synonyms[j].find(left_bracket) != -1:
            count = count + 1
            herb = str(herb_replaced_synonyms[j])
            herb = herb[1:-1]  # 截取除去【】的子串
            # 如果存在于中药别名表中，则直接替换
            if herb in herb_synonyms.keys():
                herb_replaced_synonyms[j] = herb_synonyms[herb]
                count = count - 1

    """
        如果count不为0，则表示别名替换操作后，列H还有没被替换的药材，需要保存到"别名无法替换的方子.xlsx",
        否则保存到"别名替换后的方子.xlsx"。保存的时候需要注意，将列F和列H的格式一并规格化
    """
    herb_replaced_synonyms = list_to_str(herb_replaced_synonyms)  # str数据类型
    if count == 0:
        copy_sheet(sheet, sheet1, i, k1, herb_total, herb_replaced_synonyms)
        k1 = k1 + 1
    else:
        copy_sheet(sheet, sheet2, i, k2, herb_total, herb_replaced_synonyms)
        k2 = k2 + 1

excel1.save(r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\别名替换后的方子.xlsx')
excel2.save(r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\别名无法替换的方子.xlsx')