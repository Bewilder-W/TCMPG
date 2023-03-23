import openpyxl
import re
import numpy as np
"""
    程序描述：
        	对第一阶段所做的方子的结构体一个改进，以"84463方子-10000.xlsx"表中的data
            子表为操作对象，建立一个存储结构体类，其中每一行数据代表存储为一个对象，
            类的数据成员（属性）是每一列值，保存的数据类型均为string。这个程序是可以
            直接import的，通过输入文件路径，表格名字来调用open_excel（）函数获取
            sheet对象和所有方子的对象。程序中包含standard_herbtotal（）函数，可以对
            列F进行规格化，具体规格化内容可以查看函数前的注释。并且在main函数中可以发现
            列G和列H中的药材名是一一对应的。
        
    程序输入：path（绝对路径）， sheet_name（表格名字，类型为字符串）
    程序输出：sheet（表格对象）， fangzi_object（所有方子对象的集合，类型为列表）
"""

class Struct_of_fangzi():

    def __init__(self, sheet, i):

        # 每一列都是类的一个属性
        self.id = str(sheet.cell(i + 1, 1).value)                        # 编号
        self.filename = str(sheet.cell(i + 1, 2).value)                  # 文件名
        self.fangzi_name = str(sheet.cell(i + 1, 3).value)               # 方名
        self.fangzi_synonyms = str(sheet.cell(i + 1, 4).value)           # 别名
        self.fangzi_source = str(sheet.cell(i + 1, 5).value)             # 处方来源

        self.herb_total = str(sheet.cell(i + 1, 6).value)                # 药物组成
        self.herb_non_unit = str(sheet.cell(i + 1, 7).value)             # 药物组成-herb
        self.herb_replaced_synonyms = str(sheet.cell(i + 1, 8).value)    # 【】是没有对到中药列表上的中药

        self.add_and_subtract = str(sheet.cell(i + 1, 9).value)          # 加减
        self.theory = str(sheet.cell(i + 1, 10).value)                   # 方论
        self.efficacy = str(sheet.cell(i + 1, 11).value)                 # 功效
        self.indication = str(sheet.cell(i + 1, 12).value)               # 主治
        self.prepare_method = str(sheet.cell(i + 1, 13).value)           # 制备方法
        self.usage_dosage = str(sheet.cell(i + 1, 14).value)             # 用法用量
        self.host = str(sheet.cell(i + 1, 15).value)                     # 主
        self.contraindication = str(sheet.cell(i + 1, 16).value)         # 用药禁忌
        self.clinical_application = str(sheet.cell(i + 1, 17).value)     # 临床应用
        self.pharmacology = str(sheet.cell(i + 1, 18).value)             # 药理作用
        self.discuss = str(sheet.cell(i + 1, 19).value)                  # 各家论述
        self.note = str(sheet.cell(i + 1, 20).value)                     # 备注

    # 用空格分隔药物组成这一列
    def spilt_herbtotal(self, string_of_cellvalue):
        result = string_of_cellvalue.split(' ')
        return result

    # 用空格分隔药物组成-herb这一列
    def split_herbnonunit(self, string_of_cellvalue):
        result = string_of_cellvalue.split(' ')
        return result

    # 用空格分隔【】是没有对到中药列表上的中药这一列
    def split_herbreplacedsynonyms(self, string_of_cellvalue):
        result = string_of_cellvalue.split(' ')
        return result

    # 输出属性herb_total
    def print_herbtotal(self):
        print(self.herb_total)

    # 输出属性herb_non_unit
    def print_herbnonunit(self):
        print(self.herb_non_unit)

    # 输出属性herb_replaced_synonyms
    def print_herbreplacedsynonyms(self):
        print(self.herb_replaced_synonyms)

"""
    输入药物组成一行数据,输入数据类型为字符串
    1.先删除药物组成中括号内的研制方法等描述性语句
    2.再将1中结果的标点符号规范化,统一用英文逗号隔开
    3.再将结果按照英文逗号隔开
    返回结果列表，返回数据类型为列表
"""
def standard_herbtotal(string_of_cellvalue):

    # 删除中文括号和英文括号里的描述内容
    string_of_cellvalue = delete_chinese_brackets(string_of_cellvalue)
    string_of_cellvalue = delete_english_brackets(string_of_cellvalue)

    string_of_cellvalue = re.sub(r'([^\d])\.([^\d])', r'\1 \2', string_of_cellvalue)
    string_of_cellvalue = string_of_cellvalue.replace('。', ' ')
    string_of_cellvalue = string_of_cellvalue.replace('；', ' ')
    string_of_cellvalue = string_of_cellvalue.replace('、', ' ')
    string_of_cellvalue = string_of_cellvalue.replace('，', ' ')
    string_of_cellvalue = string_of_cellvalue.replace(',', ' ')
    string_of_cellvalue = string_of_cellvalue.replace(';', ' ')
    string_of_cellvalue = string_of_cellvalue.replace('：', ' ')
    string_of_cellvalue = string_of_cellvalue.replace(':', ' ')
    string_of_cellvalue = re.sub(".$", "", string_of_cellvalue)

    result = split_column_FGH(string_of_cellvalue)
    return result

# 删除中文括号内的内容（）
def delete_chinese_brackets(string_of_cellvalue):
    left_bracket = '（'
    right_bracket = '）'

    bracket = 0  # 记录左右括号的个数
    result = ""  # 保存删除括号后的结果

    for k in range(len(string_of_cellvalue)):
        if string_of_cellvalue[k] == left_bracket:
            bracket = bracket + 1
        elif string_of_cellvalue[k] == right_bracket:
            bracket = bracket - 1
        elif bracket == 0:
            result = result + string_of_cellvalue[k]

    return result

# 删除英文括号内的内容（）
def delete_english_brackets(string_of_cellvalue):
    left_bracket = '('
    right_bracket = ')'

    bracket = 0  # 记录左右括号的个数
    result = ""  # 保存删除括号后的结果

    for k in range(len(string_of_cellvalue)):
        if string_of_cellvalue[k] == left_bracket:
            bracket = bracket + 1
        elif string_of_cellvalue[k] == right_bracket:
            bracket = bracket - 1
        elif bracket == 0:
            result = result + string_of_cellvalue[k]

    return result

# 打开路径为path的表格，返回表格对象和所有方子的对象
def open_excel(path, sheet_name):
    excel = openpyxl.load_workbook(path)
    sheet = excel[sheet_name]

    # 保存所有方子的对象
    fangzi_object = []

    for i in range(sheet.max_row):
        fangzi = Struct_of_fangzi(sheet, i)
        fangzi_object.append(fangzi)

    return sheet, fangzi_object

# 用空格分离列f，列G和列H，去除列表空值, 返回值类型为list
def split_column_FGH(string_of_cellvalue):
    result = string_of_cellvalue.split(' ')
    result = [i for i in result if i != '']
    return result

if __name__ == '__main__':
    path = r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\63702方子.xlsx'
    sheet_name = '方子'
    sheet, fangzi_object = open_excel(path, sheet_name)
    # print(sheet.max_row, len(fangzi_object))

    """
        通过此操作发现列F、列G和列H都是一一对应的
    """
    # for fangzi in fangzi_object:
    #     herb_total = fangzi.herb_total
    #     herb_total = split_column_FGH(herb_total)
    #
    #     herb_non_unit = fangzi.herb_non_unit
    #     herb_non_unit = split_column_FGH(herb_non_unit)
    #
    #     herb_replaced_synonyms = fangzi.herb_replaced_synonyms
    #     herb_replaced_synonyms = split_column_FGH(herb_replaced_synonyms)
    #
    #     if len(herb_non_unit) != len(herb_replaced_synonyms) or len(herb_total) != len(herb_non_unit):
    #         print(fangzi.id)
    #         # print(herb_total)
    #         # print(herb_non_unit)
    #         # print(herb_replaced_synonyms)

    num_eff = 0
    num_indic = 0
    num_both = 0
    for fangzi in fangzi_object:
        efficacy = fangzi.efficacy
        indication = fangzi.indication
        if indication == "-" and efficacy == "-":
            num_both += 1
        elif efficacy == "-":
            num_eff += 1
        elif indication == "-":
            num_indic += 1
    print("功效为 - 的方子数量： " + str(num_eff))
    print("主治为 - 的方子数量： " + str(num_indic))
    print("功效和主治都为 - 的方子数量： " + str(num_both))

    # outpath = r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\测试样例.txt'
    # with open(outpath, 'w', encoding="utf-8") as file:
    #     fangzi_list = np.random.randint(1, 84462, 200)
    #     fangzi_list = fangzi_list.tolist()
    #     print(type(fangzi_list))
    #     print(fangzi_list)
    #     k = 0
    #     for fangzi in fangzi_object:
    #         if k == 0:
    #             k = 1
    #             continue
    #         count = int(fangzi.id)
    #         if count in fangzi_list:
    #             print(1)
    #             file.write(fangzi.id + '          ' + fangzi.fangzi_name + '       ' + fangzi.herb_non_unit + '            ' + fangzi.indication)
    #             file.write('\n')