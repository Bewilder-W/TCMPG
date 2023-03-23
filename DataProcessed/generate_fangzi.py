import openpyxl
import numpy as np
from scipy import sparse
from gensim.models import word2vec, Word2Vec

"""
    程序描述：
        对20998方子.xlsx进行操作，得到方子—主治短语的邻接矩阵，方子的初始特征矩阵，主治短语的初始特征矩阵
"""


def w2v(indication_file_path, w2vmodel_path):
    # 保存的语料需要是一行一句话，并且词与词之间用空格隔开
    sentences = word2vec.LineSentence(indication_file_path)

    # vector_size表示输出embedding的维度,sg=1表示用skip_gram算法
    model = word2vec.Word2Vec(sentences,
                              vector_size=300,
                              window=10,
                              min_count=1,
                              seed=123,
                              workers=1,
                              sg=1,
                              sample=1e-3)

    model.save(w2vmodel_path)
    return model


def generate_p():
    # 将所有方子的主治作为主治语料库，将其放入word2vec内的skip_gram算法进行向量表示，保存构建的模型
    indication_model = Word2Vec.load('./data/indication.model')
    # feature_i = np.empty(shape=(0, 300))  # 初始生成一个空的数组，类型为ndarray,默认为float64

    path = r'./data/49493药方.xlsx'
    sheet_name = '方子'
    excel = openpyxl.load_workbook(path)
    sheet = excel[sheet_name]

    # path1 = r'C:\Users\12716\Desktop\寒假工作文件\新方子—规格化.xlsx'
    # sheet1_name = '全部'
    # excel1 = openpyxl.load_workbook(path1)
    # sheet1 = excel1[sheet1_name]

    # 将每个方子id（prescription），中药药材（herb），主治短语（indication）都用唯一的序号来表示
    index_p = 0
    index_h = 0
    index_i = 0

    # 将方子，药材，主治短语与序号之间的对应关系，用字典保存,将方子与主治短语的连接关系，也用字典表示
    dict_p = {}  # 格式为： 方子:方子的序号
    dict_h = {}  # 格式为： 中药药材:药材的序号
    dict_i = {}  # 格式为： 主治短语:主治短语的序号
    # dict_p_i = {}  # 格式为：方子的序号:主治短语的序号

    # 生成方子-主治短语的邻接矩阵
    row_p_i = []  # 记录方子-主治短语的非零元素的行，输入为方子的序号
    col_p_i = []  # 记录方子-主治短语的非零元素的列，输入为主治短语的序号
    data_p_i = []  # 记录方子-主治短语的非零元素的值，输入为1，表示有边

    # 生成每个方子的初始embedding
    row_p = []  # 输入为方子的序号
    col_p = []  # 记录每个方子的中药药材的非零元素的序号，输入为中药药材的序号
    data_p = []  # 记录每个方子对应的中药药材的值，输入为剂量占比

    # 生成每个症状的初始embedding
    row_s = []  # 输入为症状的序号
    col_s = []  # 记录每个方子的中药药材的非零元素的序号，输入为中药药材的序号
    data_s = []  # 记录每个方子对应的中药药材的值，输入为剂量占比

    maxlen_i = 0  # 统计方子含有最多数量的主治短语
    num_i = 0  # 统计所有方子的边数（方子-主治短语）

    for i in range(sheet.max_row):

        p = sheet.cell(i + 1, 1).value  # type:int
        herb_total = sheet.cell(i + 1, 3).value
        indication = sheet.cell(i + 1, 5).value

        if p not in dict_p.keys():
            dict_p[p] = index_p
            index_p += 1

        herb_total = herb_total.split(',')
        for h in herb_total:
            h = h.split(' ')

            herb = h[0].strip()  # type:str
            dose = h[1].strip()  # type:str

            if h[0] not in dict_h.keys():
                dict_h[h[0]] = index_h
                index_h += 1
            row_p.append(dict_p[p])
            col_p.append(dict_h[herb])
            # data_p.append(float(dose))
            data_p.append(float(1))

        indication = indication.split(' ')

        indication = list(set(indication))

        maxlen_i = max(len(indication), maxlen_i)
        num_i += len(indication)
        for s in indication:
            s = s.strip()
            if s not in dict_i.keys():
                dict_i[s] = index_i
                index_i += 1
                row_s.append(dict_i[s])
                col_s.append(dict_i[s])
                data_s.append(float(1))

                # feature_i = np.vstack((feature_i, indication_model.wv[s]))
            # dict_p_i[dict_p[p]] = dict_i[s]

            # 添加 i->j
            row_p_i.append(dict_p[p])
            col_p_i.append(dict_i[s])
            data_p_i.append(1)

    # print("方子总数量为：" + str(index_p))
    # print("不同中药药材总数量： " + str(index_h))
    # print("方子中主治短语的最大数量： " + str(maxlen_i))
    # print("所有方子—主治短语的边数：" + str(num_i))

    # path_matrix_p_i = r'./data/p_s_adj_matrix.npz'
    # matrix_p_i = sparse.coo_matrix((data_p_i, (row_p_i, col_p_i)), shape=(index_p, index_i))
    # sparse.save_npz(path_matrix_p_i, matrix_p_i)  # 将稀疏矩阵保存为.npz文件
    # # matrix_p_i = sparse.load_npz(path_matrix_p_i)  # 对文件路径为path_matrix_p_i的稀疏矩阵进行读取
    # print(matrix_p_i.shape)
    # print(matrix_p_i.dtype)
    # print(matrix_p_i.nnz)
    # print(matrix_p_i.toarray())

    path_feature_p = r'./data/feature_p_none_dose.npz'
    # path_feature_p = r'./data/feature_p.npz'
    feature_p = sparse.coo_matrix((data_p, (row_p, col_p)), shape=(index_p, index_h))
    sparse.save_npz(path_feature_p, feature_p)
    # feature_p = sparse.load_npz(path_feature_p)  # 对文件路径为path_feature_p的稀疏矩阵进行读取
    print(feature_p.shape)
    print(feature_p.dtype)
    print(feature_p.nnz)
    print(feature_p.toarray())

    path_feature_s = r'./data/feature_s_none_w2v.npz'
    feature_s = sparse.coo_matrix((data_s, (row_s, col_s)), shape=(index_i, index_i))
    sparse.save_npz(path_feature_s, feature_s)
    # path_feature_i = r'./data/feature_i_none_w2v.npz'
    # np.savez(path_feature_i, feature_i=np.identity(2494))
    # # 读取feature_i.npz文件：
    # path_feature_i = r'./data/feature_s.npz'
    # np.savez(path_feature_i, feature_s=feature_i)

    # w = np.load(path_feature_i)
    # feature_s = w["feature_s"]
    print(type(feature_s))
    print(feature_s.shape)
    print(feature_s.dtype)
    print(feature_s)

    # path1 = r'./data/index_herb'
    # path2 = r'./data/index_prescription'
    # path3 = r'./data/index_symptom'
    # with open(path1, 'w', encoding='utf-8') as file:
    #     for i in dict_h.keys():
    #         file.write("%5s\t %10s\n" % (str(dict_h[i]), str(i)))
    # with open(path2, 'w', encoding='utf-8') as file:
    #     for i in dict_p.keys():
    #         file.write("%10s %10s\n" % (str(dict_p[i]), str(i)))
    # with open(path3, 'w', encoding='utf-8') as file:
    #     for i in dict_i.keys():
    #         file.write("%5s %10s\n" % (str(dict_i[i]), str(i)))


if __name__ == '__main__':
    # w2v(r'./data/Indications.txt', './data/indication.model')
    generate_p()
