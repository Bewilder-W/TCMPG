import fangzi_strcut
import re
import openpyxl
"""
    目前方子中只含两种单位，g和%。将这两种单位都换算成比例，
    以 药材+比例 的格式存放
"""

path = r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\别名替换后的方子2-60.xlsx'
sheet_name = '方子'
excel = openpyxl.load_workbook(path)
sheet = excel[sheet_name]


for w in range(sheet.max_row):
    if w == 0:
        continue
    fangzi = fangzi_strcut.Struct_of_fangzi(sheet, w)

    # type: str
    # print(type(fangzi.herb_total))

    # 把这两列的str 按照空格分开， type变成list
    herb_total = fangzi_strcut.split_column_FGH(fangzi.herb_total)
    herb_non_unit = fangzi_strcut.split_column_FGH(fangzi.herb_non_unit)

    sum = 0
    for i, j in zip(herb_total, herb_non_unit):
        i = i.replace(j, '')
        if i[-1] == 'g':
            i = i.replace('g', '')
        else:
            i = i.replace('%', '')
        sum += float(i)

    res_herb_total = ""
    for i, j in zip(herb_total, herb_non_unit):
        i = i.replace(j, '')
        if i[-1] == 'g':
            i = i.replace('g', '')
        else:
            i = i.replace('%', '')

        # 控制输出的比例格式为 小数点后十位
        proportion = '%.10f'%(float(i)/sum)
        res_herb_total += j + proportion + " "

    # print(res_herb_total)
    res_herb_total = res_herb_total.rstrip()
    # # 将改完后的结果更新进表中
    sheet.cell(w + 1, 6, res_herb_total)

# 保存新的方子表
excel.save(r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\65788方子.xlsx')