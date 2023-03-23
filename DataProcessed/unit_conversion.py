import fangzi_strcut
import re
import openpyxl
"""
    根据目前存有的复合型单位，例如：两钱，钱分，两分，斤两，两钱分，钱分厘，分厘，钱厘    
    和 斤，两，钱，分，厘，字，铢 都换算成g来计算。
"""

path = r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\别名替换后的方子1-5.xlsx'
sheet_name = '方子'
excel = openpyxl.load_workbook(path)
sheet = excel[sheet_name]

for w in range(sheet.max_row):
    if w == 0:
        continue
    fangzi = fangzi_strcut.Struct_of_fangzi(sheet, w)

    # type: str
    # print(type(fangzi.herb_total))

    # 把这两列的str 按照空格分开， type变成list
    herb_total = fangzi_strcut.split_column_FGH(fangzi.herb_total)
    herb_non_unit = fangzi_strcut.split_column_FGH(fangzi.herb_non_unit)

    res_herb_total = ""
    for m, n in zip(herb_total, herb_non_unit):
        m = str(m.replace(str(n), ''))
        sums = 0
        j = 0
        if re.search(r'[斤两钱分厘字铢]', m) is not None:
            for i in range(len(m)):
                if m[i] == '斤':
                    try:
                        sums += float(m[j:i]) * 500
                    except ValueError:
                        print(str(w) + " " + m)
                    j = i + 1
                elif m[i] == '两':
                    try:
                        sums += float(m[j:i]) * 31.25
                    except ValueError:
                        print(str(w) + " " + m)
                    j = i + 1
                elif m[i] == '钱':
                    try:
                        sums += float(m[j:i]) * 3.125
                    except ValueError:
                        print(str(w) + " " + m)
                    j = i + 1
                elif m[i] == '分':
                    try:
                        sums += float(m[j:i]) * 0.3125
                    except ValueError:
                        print(str(w) + " " + m)
                    j = i + 1
                elif m[i] == '厘':
                    try:
                        sums += float(m[j:i]) * 0.03125
                    except ValueError:
                        print(str(w) + " " + m)
                    j = i + 1
                elif m[i] == '字':
                    try:
                        sums += float(m[j:i]) * 0.4
                    except ValueError:
                        print(str(w) + " " + m)
                    j = i + 1
                elif m[i] == '铢':
                    try:
                        sums += float(m[j:i]) * 0.65
                    except ValueError:
                        print(str(w) + " " + m)
                    j = i + 1
            res_herb_total += n + str(sums) + 'g '
        else:
            res_herb_total += n + m + ' '
    res_herb_total = res_herb_total.rstrip()

    # # 将改完后的结果更新进表中
    sheet.cell(w + 1, 6, res_herb_total)

    # print(fangzi.id)

# 保存新的方子表
excel.save(r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\别名替换后的方子1-6.xlsx')

