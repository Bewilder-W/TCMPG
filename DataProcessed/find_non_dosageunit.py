import fangzi_strcut
import openpyxl
"""
    程序描述：
        从别名替换后的方子1-2.xlsx中筛选得到13082个方子存在无剂量单位
        将结果保存在含有无剂量单位的方子.txt中。
"""

# 导入excel表格
# path = r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\别名替换后的方子1-4.xlsx'
path = r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\1_含有无剂量单位药材的方子-zb(1).xlsx'
sheet_name = '方子'
sheet, fangzi_object = fangzi_strcut.open_excel(path, sheet_name)

# 新建一个excel表格，保存含有无剂量单位药材的方子
excel2 = openpyxl.Workbook()
sheet2 = excel2.create_sheet('方子')



# 记录没有单位的方子
nonunit_fangzi = []  # 保存的是方子对象

for fangzi in fangzi_object:
    herb_total = fangzi_strcut.split_column_FGH(fangzi.herb_total)
    herb_non_unit = fangzi_strcut.split_column_FGH(fangzi.herb_non_unit)

    for i, j in zip(herb_total, herb_non_unit):
        k = str(i.replace(str(j), ''))
        if k == '':
            nonunit_fangzi.append(fangzi)
            break

k = 0
for fangzi in nonunit_fangzi:
    k += 1
    sheet2.cell(k, 1, fangzi.id)
    sheet2.cell(k, 2, fangzi.filename)
    sheet2.cell(k, 3, fangzi.fangzi_name)
    sheet2.cell(k, 4, fangzi.fangzi_synonyms)
    sheet2.cell(k, 5, fangzi.fangzi_source)

    sheet2.cell(k, 6, fangzi.herb_total)
    sheet2.cell(k, 7, fangzi.herb_non_unit)
    sheet2.cell(k, 8, fangzi.herb_replaced_synonyms)

    sheet2.cell(k, 9, fangzi.add_and_subtract)
    sheet2.cell(k, 10, fangzi.theory)
    sheet2.cell(k, 11, fangzi.efficacy)
    sheet2.cell(k, 12, fangzi.indication)
    sheet2.cell(k, 13, fangzi.prepare_method)
    sheet2.cell(k, 14, fangzi.usage_dosage)
    sheet2.cell(k, 15, fangzi.host)
    sheet2.cell(k, 16, fangzi.contraindication)
    sheet2.cell(k, 17, fangzi.clinical_application)
    sheet2.cell(k, 18, fangzi.pharmacology)
    sheet2.cell(k, 19, fangzi.discuss)
    sheet2.cell(k, 20, fangzi.note)


excel2.save(r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\含有无剂量单位药材的方子.xlsx')