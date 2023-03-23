import fangzi_strcut
import re
import openpyxl
"""
    将其他的单位 换算成 g
"""

path = r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\别名替换后的方子2-59.xlsx'
sheet_name = '方子'
excel = openpyxl.load_workbook(path)
sheet = excel[sheet_name]

for w in range(sheet.max_row):
    if w == 0:
        continue
    fangzi = fangzi_strcut.Struct_of_fangzi(sheet, w)

    # type: str
    # print(type(fangzi.herb_total))

    # 把这两列的str 按照空格分开， type变成list
    herb_total = fangzi_strcut.split_column_FGH(fangzi.herb_total)
    herb_non_unit = fangzi_strcut.split_column_FGH(fangzi.herb_non_unit)

    res_herb_total = ""
    for m, n in zip(herb_total, herb_non_unit):
        m = str(m.replace(str(n), ''))
        sums = 0
        j = 0
        if re.search(r'[升合尺寸]', m) is not None:
            for i in range(len(m)):
                if m[i] == '升':
                    try:
                        sums += float(m[j:i]) * 400
                    except ValueError:
                        print(str(w) + " " + m)
                    j = i + 1
                elif m[i] == '合':
                    try:
                        sums += float(m[j:i]) * 40
                    except ValueError:
                        print(str(w) + " " + m)
                    j = i + 1
                elif m[i] == '尺':
                    try:
                        sums += float(m[j:i]) * 15
                    except ValueError:
                        print(str(w) + " " + m)
                    j = i + 1
                elif m[i] == '寸':
                    try:
                        sums += float(m[j:i]) * 1.5
                    except ValueError:
                        print(str(w) + " " + m)
                    j = i + 1
            res_herb_total += n + str(sums) + 'g '
        else:
            res_herb_total += n + m + ' '
    res_herb_total = res_herb_total.rstrip()

    # # 将改完后的结果更新进表中
    sheet.cell(w + 1, 6, res_herb_total)

    # res_herb_total = ""
    #
    # # 修改不同的单位 和 单位质量（换算成g）
    # unit = "鸡子大"
    # unit_dosage = 30
    #
    # for m, n in zip(herb_total, herb_non_unit):
    #     m = str(m.replace(str(n), ''))
    #     org_m = m
    #     #     print(m+org_m)
    #
    #     # 取到浮点数或者整数  取到的结果是list类型
    #     result = re.findall(r'\d+\.\d+|\d+', m)
    #     #     print(result)
    #     #     print(type(result))
    #     if len(result) != 0:
    #         result = result[0]  # 转换成str
    #
    #     # 删除数字和小数点后，得到的就是单位
    #     m = re.sub('\d+', '', m)
    #     m = re.sub('\.', '', m)
    #
    #     #     print(m)
    #     if m == unit:
    #         sums = float(result) * unit_dosage
    #         res_herb_total += n + str(sums) + 'g '
    #     else:
    #         res_herb_total += n + org_m + ' '
    #
    # res_herb_total = res_herb_total.rstrip()
    #
    # # # 将改完后的结果更新进表中
    # sheet.cell(w + 1, 6, res_herb_total)

    # print(fangzi.id)

# 保存新的方子表
excel.save(r'C:\Users\吴杨\Desktop\中药复方-吴杨\TCM\别名替换后的方子2-60.xlsx')

